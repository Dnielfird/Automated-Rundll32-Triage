import os
import sys
import re
import argparse
from datetime import datetime, timedelta
from dateutil import parser as date_parser
from dateutil.relativedelta import relativedelta
from dotenv import load_dotenv
from elasticsearch import Elasticsearch
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

# --- CONFIGURATION ---
load_dotenv()

ES_URL = os.getenv("ES_URL")
ES_API_KEY = os.getenv("ES_API_KEY")
ES_INDEX = os.getenv("ES_INDEX", "*")
VERIFY_CERTS = os.getenv("VERIFY_CERTS", "false").lower() == "true"

TARGET_PROCESS = "rundll32.exe"

# --- COLORS ---
COLOR_TP_RED   = "F4C7C3"  # Red for Suspicious/Malicious
COLOR_HEADER   = "1F4E78"  # Dark Blue for Headers

# --- FIELD LIST ---
FIELDS = [
    "@timestamp", 
    "host.name",
    "user.name",
    "event.action",
    "event.dataset",
    "user.domain",
    "event.code",
    "process.name",
    "process.command_line",
    "process.args",
    "process.working_directory",
    "process.entity_id",
    "process.parent.name",
    "process.parent.command_line",
    "process.parent.executable",
    "destination.ip",
    "destination.port",
    "dns.question.name",
    "file.path",
    "file.target_path",
    "registry.path",
    "process.code_signature.exists",
    "process.code_signature.status",
    "process.code_signature.subject_name",
    "process.code_signature.trusted",
    "process.pe.original_file_name",
    "process.hash.sha256",
    "dll.code_signature.exists",
    "dll.name",
    "dll.path",
    "dll.hash.sha256"
]

# ==========================================
#       THREAT ANALYZER ENGINE
# ==========================================

class ThreatAnalyzer:
    def __init__(self):
        pass

    # ---------------------------------------------------------
    # GLOBAL NOISE FILTER (UPDATED FROM YOUR IMAGES)
    # ---------------------------------------------------------
    def is_known_noise(self, cmd, parent):
        """
        Filters out known high-volume system noise.
        """
        cmd_lower = cmd.lower()
        parent_lower = parent.lower()

        if "ndfapi.dll" in cmd_lower and "ndfrundlldiagnose" in cmd_lower:
            return True, "Noise: Windows Network Diagnostics"

        if "pcasvc.dll" in cmd_lower and "pcapatchsdbtask" in cmd_lower:
            return True, "Noise: Program Compatibility Assistant"

        if "edgehtml.dll" in cmd_lower and "#141" in cmd_lower:
            return True, "Noise: Edge/Store App Background Task"

        if "startupscan.dll" in cmd_lower and "susruntask" in cmd_lower:
             return True, "Noise: Windows Startup Scan"

        if "iesetup.dll" in cmd_lower and "iehardenuser" in cmd_lower:
             return True, "Noise: IE Security Config"
        
        if "windows.staterepositoryclient.dll" in cmd_lower:
             return True, "Noise: Windows State Repository"

        if "shcreatelocalserverrundll" in cmd_lower and any(p in parent_lower for p in ["svchost", "explorer", "runtimebroker"]):
             return True, "Noise: Safe DCOM/COM Server Hosting"

        if "cryptext.dll" in cmd_lower and "cryptextaddcer" in cmd_lower:
             return True, "Noise: Certificate Installation"

        if "acproxy.dll" in cmd_lower:
             return True, "Noise: Autocheck Proxy"

        if "shell32.dll" in cmd_lower:
            if "control_rundll" in cmd_lower and any(cpl in cmd_lower for cpl in ["sysdm.cpl", "timedate.cpl", "inetcpl.cpl", "mmsys.cpl", "ncpa.cpl", "desk.cpl"]):
                return True, "Noise: Standard Control Panel"
            if "openas_rundll" in cmd_lower:
                 return True, "Noise: 'Open With' Dialog"

        if "inetcpl.cpl" in cmd_lower and "clearmytracksbyprocess" in cmd_lower:
            return True, "Noise: Browser History Cleanup"

        if any(x in cmd_lower for x in ["nvcontainer", "nvidia", "amd software", "radeon", "intel graphics"]):
            return True, "Noise: Graphics Driver"

        return False, ""

    # ---------------------------------------------------------
    # RULE 1: CREDENTIAL DUMPING
    # ---------------------------------------------------------
    def _check_rule_1_creds(self, cmd, parent):
        if "comsvcs.dll" in cmd and "minidump" in cmd:
            is_tp = True
            reason = "LSASS Dump Detected. CAUTION: Verify Parent is NOT SCCM/Tanium."
            if any(x in parent for x in ["ccmexec", "tanium", "sccm"]): 
                is_tp = False; reason = "Legitimate Admin Tool (SCCM/Tanium)"
            return {'is_hit': True, 'is_tp': is_tp, 'reason': reason}
        return {'is_hit': False, 'is_tp': False, 'reason': ""}

    # ---------------------------------------------------------
    # RULE 2: MASQUERADING
    # ---------------------------------------------------------
    def _check_rule_2_masq(self, cmd, parent):
        if re.search(r'\.(txt|jpg|png|dat|tmp)$', cmd) or re.search(r'\.(txt|jpg|png|dat|tmp)\s*,', cmd):
            is_tp = True
            reason = "Non-DLL Extension. CAUTION: Check File Header (MZ)."
            
            if ".dat" in cmd and "system32" in cmd:
                is_tp = False; reason = "Legacy System .DAT (Safe)"
            elif not any(p in parent for p in ["chrome", "edge", "firefox", "word", "excel", "powershell", "cmd", "wscript"]):
                 is_tp = False; reason = "Non-DLL loaded by benign parent (Likely Safe)"
            
            return {'is_hit': True, 'is_tp': is_tp, 'reason': reason}
        return {'is_hit': False, 'is_tp': False, 'reason': ""}

    # ---------------------------------------------------------
    # RULE 3: ORDINALS
    # ---------------------------------------------------------
    def _check_rule_3_ordinal(self, cmd, parent):
        if re.search(r',\s*#', cmd):
            is_tp = True
            reason = "Ordinal Execution. CAUTION: Check Digital Signature."
            
            if any(x in cmd for x in ["nv", "nvidia", "intel", "hp", "canon", "print", "brilliance", "amd", "radeon"]):
                is_tp = False; reason = "Vendor Driver Ordinal (Safe)"
            elif "edge" in cmd or "chrome" in cmd:
                is_tp = False; reason = "Browser Helper Ordinal (Safe)"
            
            if any(path in cmd for path in ["appdata", "temp", "public", "programdata"]) and is_tp == False:
                 is_tp = True; reason = "Ordinal call from Suspicious Folder (High Risk)"
            
            return {'is_hit': True, 'is_tp': is_tp, 'reason': reason}
        return {'is_hit': False, 'is_tp': False, 'reason': ""}

    # ---------------------------------------------------------
    # RULE 4: SCRIPTING
    # ---------------------------------------------------------
    def _check_rule_4_scripting(self, cmd, parent):
        if any(x in cmd for x in ["javascript:", "vbscript:", "script:", "about:", "mshtml,runhtmlapplication"]):
            is_tp = True
            reason = "Script Execution. CAUTION: Check for 'ActiveX' or 'Eval'."
            
            if any(x in cmd for x in ["window.close", "void(0)", "print", "history.back", "window.external"]):
                is_tp = False; reason = "Benign UI Script (Safe)"
            elif any(x in parent for x in ["hh.exe", "msiexec", "unins", "setup.exe", "isbew64.exe"]):
                 if not any(bad in cmd for bad in ["wscript.shell", "powershell", "cmd.exe", "eval"]):
                    is_tp = False; reason = "Help/Installer UI Action (Safe)"
            
            if any(x in cmd for x in ["activexobject", "wscript.shell", "getobject", "eval(", "runhtmlapplication"]):
                is_tp = True; reason = "Malicious Scripting Object Detected"
            
            return {'is_hit': True, 'is_tp': is_tp, 'reason': reason}
        return {'is_hit': False, 'is_tp': False, 'reason': ""}

    # ---------------------------------------------------------
    # RULE 5: REMOTE LOADING (SMB/WebDAV)
    # ---------------------------------------------------------
    def _check_rule_5_remote_load(self, cmd, parent):
        if "\\\\" in cmd or "http:" in cmd or "ftp:" in cmd:
            is_tp = True
            reason = "Remote Loading Detected."

            if "http:" in cmd or "ftp:" in cmd:
                is_tp = True; reason = "WebDAV/Internet DLL Load (High Confidence Malicious)"
            elif "\\c$" in cmd or "\\admin$" in cmd or "\\ipc$" in cmd:
                if any(p in parent for p in ["ccmexec", "tanium", "sccm"]):
                    is_tp = False; reason = "Internal Admin Tool (SCCM/Tanium)"
                else:
                    is_tp = True; reason = "Hidden Admin Share Access (Lateral Movement)"
            elif "sysvol" in cmd or "netlogon" in cmd:
                is_tp = False; reason = "Domain Login Script (Safe)"
            else:
                reason = "Network Share Load. CAUTION: Verify Source Server."
            
            return {'is_hit': True, 'is_tp': is_tp, 'reason': reason}
        return {'is_hit': False, 'is_tp': False, 'reason': ""}

    # ---------------------------------------------------------
    # RULE 6: ADVANCED LIBRARIES
    # ---------------------------------------------------------
    def _check_rule_6_advanced_libs(self, cmd, parent):
        if any(lib in cmd for lib in ["advpack", "ieframe", "shdocvw", "syssetup", "setupapi", "url.dll", "fileprotocolhandler", "registerocx"]):
            is_tp = True
            reason = "Advanced Installer Abuse. CAUTION: Verify .inf file path."

            if "launchinfsection" in cmd and ",,1" in cmd:
                if any(path in cmd for path in ["temp", "public", "appdata", "downloads"]):
                    return {'is_hit': True, 'is_tp': True, 'reason': "Silent .INF Execution from Temp (High Confidence)"}

            if "fileprotocolhandler" in cmd:
                if any(ext in cmd for ext in [".hta", ".sct", ".vbs", ".js", ".exe", ".scr", ".pif"]):
                    is_tp = True; reason = "FileProtocolHandler launching Script/Executable"
                elif any(ext in cmd for ext in [".htm", ".pdf", ".txt", ".jpg"]) or "http" in cmd:
                     is_tp = False; reason = "Standard URL/File Open (Likely Safe)"

            if "c:\\windows\\inf" in cmd or "program files" in cmd:
                 is_tp = False; reason = "System Driver Installation (Safe)"
            elif any(p in parent for p in ["setup.exe", "msiexec.exe", "trustedinstaller.exe"]):
                 is_tp = False; reason = "Trusted Installer Parent (Safe)"

            return {'is_hit': True, 'is_tp': is_tp, 'reason': reason}
        return {'is_hit': False, 'is_tp': False, 'reason': ""}

    # ---------------------------------------------------------
    # RULE 7: EVASION (ADS, GUID, Zipfldr)
    # ---------------------------------------------------------
    def _check_rule_7_evasion(self, cmd, parent, original_cmd):
        guid_match = re.search(r'\{[0-9a-fA-F]{8}-[0-9a-fA-F]{4}-[0-9a-fA-F]{4}-[0-9a-fA-F]{4}-[0-9a-fA-F]{12}\}', original_cmd)
        hit = False; tp = False; reason = ""

        if ":" in cmd and not re.search(r'^[a-z]:\\', cmd) and "http:" not in cmd and "https:" not in cmd:
            hit = True
            if "zone.identifier" in cmd: tp = False; reason = "Zone.Identifier Metadata (Safe)"
            else: tp = True; reason = "ADS Execution (Hidden Code). CAUTION: Verify Stream."
        
        elif "zipfldr" in cmd or "routethecall" in cmd:
            hit = True; tp = True; reason = "Zipfldr Proxy. CAUTION: Check target."
            
        elif "-sta" in cmd:
            hit = True; tp = True; reason = "STA Mode. CAUTION: Check COM Object."

        elif guid_match:
            hit = True
            # The Global Filter at the top already removed the safe GUIDs from svchost.
            # If it's still here, it means the parent wasn't Svchost/Explorer, so it's suspicious.
            if any(p in parent for p in ["explorer", "setup", "msiexec"]): 
                 tp = False; reason = "Explorer GUID Load (Low Risk)"
            else: 
                tp = True; reason = "Obfuscated GUID from Suspicious Parent (High Risk)"

        if hit: return {'is_hit': True, 'is_tp': tp, 'reason': reason}
        return {'is_hit': False, 'is_tp': False, 'reason': ""}

    # ---------------------------------------------------------
    # MAIN ANALYSIS LOOP
    # ---------------------------------------------------------
    def analyze_event(self, cmd, parent, original_cmd):
        results = {}
        cmd_lower = cmd.lower() if cmd else ""
        parent_lower = parent.lower() if parent else ""

        # 1. Global Noise Check (Crucial for filtering your screenshot events)
        is_noise, noise_reason = self.is_known_noise(cmd_lower, parent_lower)
        if is_noise:
            return {'GLOBAL_FILTER': {'is_hit': False, 'is_tp': False, 'reason': noise_reason}}

        # 2. Run Each Rule (1-7)
        results['RULE_1'] = self._check_rule_1_creds(cmd_lower, parent_lower)
        results['RULE_2'] = self._check_rule_2_masq(cmd_lower, parent_lower)
        results['RULE_3'] = self._check_rule_3_ordinal(cmd_lower, parent_lower)
        results['RULE_4'] = self._check_rule_4_scripting(cmd_lower, parent_lower)
        results['RULE_5'] = self._check_rule_5_remote_load(cmd_lower, parent_lower)
        results['RULE_6'] = self._check_rule_6_advanced_libs(cmd_lower, parent_lower)
        results['RULE_7'] = self._check_rule_7_evasion(cmd_lower, parent_lower, original_cmd)

        return results

# ==========================================
#          UTILITIES (Same as before)
# ==========================================

def get_es_client():
    if not ES_URL or not ES_API_KEY: sys.exit("[!] Error: Missing ES_URL or ES_API_KEY.")
    return Elasticsearch([ES_URL], api_key=ES_API_KEY, verify_certs=VERIFY_CERTS)

def get_nested_value(doc, path):
    keys = path.split('.')
    val = doc.get('_source', {})
    try:
        for key in keys:
            val = val.get(key); 
            if val is None: return ""
        if isinstance(val, list): return " ".join([str(v) for v in val])
        return str(val)
    except: return ""

def format_sheet_header(ws, headers, color):
    ws.append(headers)
    fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
    font = Font(bold=True, color="FFFFFF")
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[get_column_letter(col_num)].width = 25

def write_colored_row(ws, row_data, is_tp):
    ws.append(row_data)
    curr_row = ws.max_row
    color = COLOR_TP_RED 
    fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
    for col in range(1, len(row_data) + 1):
        ws.cell(row=curr_row, column=col).fill = fill

def parse_time_window(t_arg):
    now = datetime.now()
    t_arg = t_arg.strip()
    if " to " in t_arg:
        try:
            start_str, end_str = t_arg.split(" to ")
            s_dt = date_parser.parse(start_str.strip(), dayfirst=True)
            e_dt = date_parser.parse(end_str.strip(), dayfirst=True)
            
            if s_dt.hour == 0 and s_dt.minute == 0 and s_dt.second == 0:
                s_dt = s_dt.replace(hour=0, minute=0, second=0)
            if e_dt.hour == 0 and e_dt.minute == 0 and e_dt.second == 0:
                e_dt = e_dt.replace(hour=23, minute=59, second=59)

            return {"gte": s_dt.isoformat(), "lte": e_dt.isoformat()}, s_dt, e_dt
        except Exception as e:
            print(f"[!] Date Parse Error: {e}. Falling back to 24h.")

    match = re.match(r"^(\d+)([hdm])$", t_arg.lower())
    if match:
        val = int(match.group(1)); unit = match.group(2)
        if unit == 'h': delta = timedelta(hours=val); es_unit = 'h'
        elif unit == 'd': delta = timedelta(days=val); es_unit = 'd'
        elif unit == 'm': delta = relativedelta(months=val); es_unit = 'M'
        return {"gte": f"now-{val}{es_unit}", "lte": "now"}, now - delta, now
    return {"gte": "now-24h", "lte": "now"}, now - timedelta(hours=24), now

def fetch_logs(es, es_range):
    print(f"[*] Querying Elasticsearch for {TARGET_PROCESS}...")
    query = {
        "bool": {
            "must": [
                {"term": {"host.os.type": "windows"}},
                {"term": {"event.type": "start"}},
                {"match": {"process.name": TARGET_PROCESS}}
            ],
            "filter": [{"range": {"@timestamp": es_range}}]
        }
    }
    data = []
    try: resp = es.search(index=ES_INDEX, body={"query": query}, size=5000, scroll='2m')
    except Exception as e: print(e); return []
    scroll_id = resp.get('_scroll_id'); hits = resp['hits']['hits']
    data.extend(hits)
    while len(hits) > 0:
        try:
            resp = es.scroll(scroll_id=scroll_id, scroll='2m'); hits = resp['hits']['hits']; data.extend(hits)
        except: break
    try: es.clear_scroll(scroll_id=scroll_id)
    except: pass
    return data

# ==========================================
#          MAIN EXECUTION
# ==========================================

def save_analysis(events, filename):
    if not events: print("[!] No events found."); return

    analyzer = ThreatAnalyzer()
    wb = Workbook(); 
    if "Sheet" in wb.sheetnames: del wb["Sheet"]

    rules = ["RULE_1", "RULE_2", "RULE_3", "RULE_4", "RULE_5", "RULE_6", "RULE_7"]
    sheets = {}

    for r in rules:
        det_sheet_name = f"{r}_Detections"
        ws_det = wb.create_sheet(det_sheet_name)
        format_sheet_header(ws_det, [f.upper() for f in FIELDS], COLOR_HEADER)
        
        val_sheet_name = f"{r}_Analysis"
        ws_val = wb.create_sheet(val_sheet_name)
        val_headers = ["STATUS", "ANALYST_CAUTION_CHECKLIST"] + [f.upper() for f in FIELDS]
        format_sheet_header(ws_val, val_headers, COLOR_HEADER)

        sheets[r] = {'det': ws_det, 'val': ws_val}

    print("[*] Analyzing events and filtering noise...")

    for event in events:
        cmd_line = get_nested_value(event, "process.command_line")
        parent_proc = get_nested_value(event, "process.parent.name")
        row_data = [get_nested_value(event, f) for f in FIELDS]

        analysis_results = analyzer.analyze_event(cmd_line, parent_proc, cmd_line)

        for rule_id, res in analysis_results.items():
            if rule_id == 'GLOBAL_FILTER': continue 

            if res['is_hit']:
                sheets[rule_id]['det'].append(row_data)
                if res['is_tp']:
                    val_row = ["TRUE POSITIVE", res['reason']] + row_data
                    write_colored_row(sheets[rule_id]['val'], val_row, True)

    wb.save(filename)
    print(f"[+] Analysis Saved to {filename}")
    print(f"[+] Noise Filters Updated: Handled Startupscan, IEsetup, StateRepository, and Safe GUIDs.")

def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("-t", "--time", default="24h", help="Time frame")
    parser.add_argument("-o", "--output", required=True, help="Output filename")
    args = parser.parse_args()

    es = get_es_client()
    es_range, _, _ = parse_time_window(args.time)
    data = fetch_logs(es, es_range)
    save_analysis(data, args.output)

if __name__ == "__main__":
    main()